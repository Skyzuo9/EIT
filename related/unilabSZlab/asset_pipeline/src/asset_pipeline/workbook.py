from __future__ import annotations

import csv
import hashlib
import json
import re
from pathlib import Path
from zipfile import ZipFile
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook

from .models import DeviceRecord, WorkflowStatus
from .state import StateStore


COVERAGE_AMBIGUITY_MARKERS = (
    "型号未注明",
    "厂商未注明",
    "自制",
    "XPR/XSR",
    "9300/9320",
    "Pro / Multitron",
    "UP20/UP20L",
    "GEN1/GEN2",
    "BL-1300 / BL-600HA",
)

TEAM_INPUT_HEADERS = {
    "name": ("仪器名称", "设备名称"),
    "official_url": ("官网或产品页", "官网链接", "产品页"),
    "width": ("宽度mm", "宽mm"),
    "depth": ("深度mm", "深mm"),
    "height": ("高度mm", "高mm"),
    "dimension_url": ("尺寸来源链接", "尺寸来源URL"),
    "notes": ("用途或备注", "备注"),
}


HEADERS = {
    "A": "记录类型",
    "B": "设备组",
    "C": "设备类型",
    "D": "厂商与型号",
    "E": "仿真资产准备状况",
    "G": "STL|Xacro状态",
    "M": "参数",
    "S": "模型可用性",
    "T": "模型证据",
    "U": "模型仓库链接",
    "V": "官网|资料链接",
    "W": "模型适配建议",
    "Y": "结构化尺寸信息",
}
MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def _column_name(reference: str) -> str:
    return "".join(char for char in reference if char.isalpha())


def _read_ooxml_rows(path: Path) -> list[dict[str, str]]:
    with ZipFile(path) as archive:
        worksheet_names = [
            name
            for name in archive.namelist()
            if name.startswith("xl/worksheets/sheet")
        ]
        if not worksheet_names:
            raise ValueError(f"No worksheet XML found in {path}")
        worksheet_name = max(
            worksheet_names, key=lambda name: archive.getinfo(name).file_size
        )
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            shared_root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in shared_root.findall(f"{{{MAIN_NS}}}si"):
                shared_strings.append("".join(item.itertext()))
        root = ET.fromstring(archive.read(worksheet_name))

    rows: list[dict[str, str]] = []
    for row in root.findall(f".//{{{MAIN_NS}}}sheetData/{{{MAIN_NS}}}row"):
        values: dict[str, str] = {}
        for cell in row.findall(f"{{{MAIN_NS}}}c"):
            column = _column_name(cell.attrib.get("r", ""))
            inline = cell.find(f"{{{MAIN_NS}}}is")
            value = cell.find(f"{{{MAIN_NS}}}v")
            if inline is not None:
                text = "".join(inline.itertext())
            elif value is not None and value.text is not None:
                text = value.text
                if cell.attrib.get("t") == "s":
                    text = shared_strings[int(text)]
            else:
                text = ""
            values[column] = text
        rows.append(values)
    return rows


def read_rows(path: Path) -> list[dict[str, str]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        if sheet.max_row > 1 and sheet.max_column > 1:
            rows: list[dict[str, str]] = []
            for values in sheet.iter_rows(values_only=True):
                rows.append(
                    {
                        _number_to_column(index): "" if value is None else str(value)
                        for index, value in enumerate(values, start=1)
                    }
                )
            return rows
    except Exception:
        pass
    return _read_ooxml_rows(path)


def _number_to_column(number: int) -> str:
    result = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        result = chr(65 + remainder) + result
    return result


def stable_device_id(name: str, source_row: int) -> str:
    slug = re.sub(r"[^\w]+", "-", name.casefold(), flags=re.UNICODE).strip("-")
    slug = slug[:56] or "device"
    digest = hashlib.sha1(f"{source_row}:{name}".encode()).hexdigest()[:8]
    return f"{slug}-{digest}"


def stable_source_device_id(source_key: str, name: str) -> str:
    slug = re.sub(r"[^\w]+", "-", source_key.casefold(), flags=re.UNICODE).strip("-")
    slug = slug[:48] or "coverage-device"
    digest = hashlib.sha1(f"{source_key}:{name}".encode()).hexdigest()[:8]
    return f"coverage-{slug}-{digest}"


def _split_links(value: str) -> list[str]:
    return [link.strip() for link in re.split(r"[,，\n]", value) if link.strip()]


def _find_team_input_columns(rows: list[dict[str, str]]) -> dict[str, str] | None:
    if not rows:
        return None
    header = rows[0]
    columns: dict[str, str] = {}
    for field, markers in TEAM_INPUT_HEADERS.items():
        for column, raw_value in header.items():
            value = str(raw_value).replace(" ", "").casefold()
            matched = (
                any(value.startswith(marker.casefold()) for marker in markers)
                if field == "name"
                else any(marker.casefold() in value for marker in markers)
            )
            if matched:
                columns[field] = column
                break
    required = {"name", "official_url", "width", "depth", "height"}
    return columns if required.issubset(columns) else None


def _team_input_value(row: dict[str, str], columns: dict[str, str], field: str) -> str:
    column = columns.get(field)
    return str(row.get(column, "")).strip() if column else ""


def write_team_instrument_names(
    path: Path,
    instrument_names: list[str],
    replace: bool = False,
) -> int:
    names = list(
        dict.fromkeys(
            " ".join(str(name).split())
            for name in instrument_names
            if str(name).strip()
        )
    )
    if not names:
        raise ValueError("Provide at least one non-empty instrument name")
    if len(names) > 30:
        raise ValueError("The team input workbook supports at most 30 instruments")

    rows = read_rows(path)
    columns = _find_team_input_columns(rows)
    if not columns:
        raise ValueError(f"Not a team input workbook: {path}")

    workbook = load_workbook(path)
    sheet = (
        workbook["待生成仪器"]
        if "待生成仪器" in workbook.sheetnames
        else workbook.active
    )
    name_column = _column_index(columns["name"])
    existing_rows = [
        row_number
        for row_number in range(2, sheet.max_row + 1)
        if str(sheet.cell(row=row_number, column=name_column).value or "").strip()
    ]
    if existing_rows and not replace:
        raise ValueError(
            "The team input workbook already contains instruments; use replace=True "
            "only before pipeline state exists"
        )
    if replace:
        for row_number in range(2, sheet.max_row + 1):
            for column_number in range(1, 8):
                sheet.cell(row=row_number, column=column_number).value = None

    for offset, name in enumerate(names, start=2):
        sheet.cell(row=offset, column=name_column).value = name
    workbook.save(path)
    return len(names)


def _import_team_input_devices(
    path: Path,
    rows: list[dict[str, str]],
    columns: dict[str, str],
) -> list[DeviceRecord]:
    devices: list[DeviceRecord] = []
    for source_row, row in enumerate(rows[1:], start=2):
        name = _team_input_value(row, columns, "name")
        if not name:
            continue
        official_url = _team_input_value(row, columns, "official_url")
        dimension_url = _team_input_value(row, columns, "dimension_url")
        notes = _team_input_value(row, columns, "notes")
        dimensions = {
            "宽mm": _team_input_value(row, columns, "width"),
            "深mm": _team_input_value(row, columns, "depth"),
            "高mm": _team_input_value(row, columns, "height"),
            "尺寸来源URL": dimension_url,
            "占地/安装备注": notes,
        }
        dimensions = {key: value for key, value in dimensions.items() if value}
        links = list(
            dict.fromkeys(_split_links(official_url) + _split_links(dimension_url))
        )
        devices.append(
            DeviceRecord(
                id=stable_device_id(name, source_row),
                source_row=source_row,
                source_kind="team_input",
                source_path=str(path.resolve()),
                source_key=f"row-{source_row}",
                record_type="实验室仪器",
                device_type="待生成3D资产",
                manufacturer_model=name,
                preparation_status="准备中",
                model_status="待生成",
                parameters=notes,
                model_availability="缺少模型资产",
                model_evidence=dimension_url,
                official_links=links,
                adaptation_advice=notes,
                structured_dimensions=(
                    json.dumps(dimensions, ensure_ascii=False) if dimensions else ""
                ),
                route="needs_generation",
            )
        )
    return devices


def import_devices(path: Path) -> tuple[list[DeviceRecord], list[dict[str, str]]]:
    rows = read_rows(path)
    if not rows:
        raise ValueError(f"Workbook has no rows: {path}")
    team_columns = _find_team_input_columns(rows)
    if team_columns:
        return _import_team_input_devices(path, rows, team_columns), rows
    devices: list[DeviceRecord] = []
    for source_row, row in enumerate(rows[1:], start=2):
        if row.get("E", "").strip() != "准备中":
            continue
        name = row.get("D", "").strip() or row.get("A", "").strip()
        model_status = row.get("G", "").strip()
        repository_link = row.get("U", "").strip()
        has_existing_mesh = (
            "Mesh可用" in model_status
            or "URDF可用" in model_status
            or bool(repository_link)
        )
        devices.append(
            DeviceRecord(
                id=stable_device_id(name, source_row),
                source_row=source_row,
                source_kind="workbook",
                source_path=str(path.resolve()),
                record_type=row.get("A", "").strip(),
                device_group=row.get("B", "").strip(),
                device_type=row.get("C", "").strip(),
                manufacturer_model=name,
                preparation_status=row.get("E", "").strip(),
                model_status=model_status,
                parameters=row.get("M", "").strip(),
                model_availability=row.get("S", "").strip(),
                model_evidence=row.get("T", "").strip(),
                repository_link=repository_link,
                official_links=_split_links(row.get("V", "")),
                adaptation_advice=row.get("W", "").strip(),
                structured_dimensions=row.get("Y", "").strip(),
                route="reuse_existing" if has_existing_mesh else "needs_generation",
            )
        )
    return devices, rows


def coverage_ambiguity_reason(name: str) -> str:
    for marker in COVERAGE_AMBIGUITY_MARKERS:
        if marker in name:
            return f"型号不唯一或不可公开检索：{marker}"
    return ""


def import_coverage_csv(path: Path) -> tuple[list[DeviceRecord], list[dict[str, str]]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    devices: list[DeviceRecord] = []
    for source_row, row in enumerate(rows, start=2):
        if row.get("模型位置", "").strip():
            continue
        name = row.get("厂商和型号", "").strip()
        source_key = row.get("注册表设备ID", "").strip()
        if not name or not source_key:
            continue
        ambiguity = coverage_ambiguity_reason(name)
        devices.append(
            DeviceRecord(
                id=stable_source_device_id(source_key, name),
                source_row=source_row,
                source_kind="coverage_csv",
                source_path=str(path.resolve()),
                source_key=source_key,
                record_type=row.get("设备大类", "").strip(),
                device_group=row.get("设备类型", "").strip(),
                device_type=row.get("设备描述", "").strip(),
                manufacturer_model=name,
                preparation_status="模型位置为空",
                model_status="待补充" if ambiguity else "待生成",
                model_availability="缺少模型资产",
                model_evidence=ambiguity,
                repository_link="",
                official_links=_split_links(row.get("oss链接", "")),
                adaptation_advice=ambiguity,
                structured_dimensions="",
                route="manual_identification" if ambiguity else "needs_generation",
            )
        )
    return devices, rows


def write_coverage_results(
    source_path: Path,
    output_path: Path,
    store: StateStore,
) -> Path:
    with source_path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = list(reader.fieldnames or [])
        rows = list(reader)
    required = {"注册表设备ID", "模型ID", "模型位置", "模型类型"}
    if not required.issubset(fieldnames):
        raise ValueError(
            f"Coverage CSV is missing columns: {sorted(required - set(fieldnames))}"
        )

    prefixes = []
    for row in rows:
        location = row.get("模型位置", "").strip()
        marker = "/模型资产/"
        if marker in location:
            prefixes.append(location.split(marker, 1)[0] + marker.rstrip("/"))
    prefix = max(set(prefixes), key=prefixes.count) if prefixes else "模型资产"

    approved = {
        device.source_key: device
        for device, status in store.list_devices()
        if device.source_kind == "coverage_csv"
        and device.source_key
        and status == WorkflowStatus.APPROVED
    }
    for row in rows:
        key = row.get("注册表设备ID", "").strip()
        device = approved.get(key)
        if not device:
            continue
        published = store.get_metadata(f"published:{device.id}")
        if not published or not Path(published).exists():
            continue
        row["模型ID"] = device.source_key
        row["模型位置"] = f"{prefix}/{Path(published).name}"
        row["模型类型"] = "gltf"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
    return output_path


def parse_dimensions_json(value: str) -> dict:
    if not value:
        return {}
    try:
        return json.loads(value)
    except json.JSONDecodeError:
        return {}


def write_results(source_path: Path, output_path: Path, store: StateStore) -> Path:
    _, source_rows = import_devices(source_path)
    devices = {
        device.source_row: (device, status)
        for device, status in store.list_devices()
        if device.source_kind == "workbook"
    }

    output = Workbook()
    sheet = output.active
    sheet.title = "数据表1"
    original_columns = max(
        (
            max((_column_index(column) for column in row), default=0)
            for row in source_rows
        ),
        default=0,
    )
    extra_headers = [
        "流水线设备ID",
        "流水线状态",
        "最终GLB路径",
        "Meshy任务ID",
        "Meshy实际Credits",
        "自动质检结论",
    ]
    for row_number, row in enumerate(source_rows, start=1):
        for column, value in row.items():
            sheet.cell(row=row_number, column=_column_index(column), value=value)
    for offset, header in enumerate(extra_headers, start=1):
        sheet.cell(row=1, column=original_columns + offset, value=header)

    for source_row, (device, status) in devices.items():
        task = store.get_meshy_task(device.id)
        qc = store.get_qc(device.id)
        published_path = store.get_metadata(f"published:{device.id}")
        values = [
            device.id,
            status.value,
            published_path or (qc.final_glb if qc else ""),
            task.task_id if task else "",
            task.consumed_credits if task else 0,
            ("通过" if qc and qc.passed else "未通过") if qc else "",
        ]
        for offset, value in enumerate(values, start=1):
            sheet.cell(row=source_row, column=original_columns + offset, value=value)

    sheet.freeze_panes = "A2"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output.save(output_path)
    return output_path


def _column_index(column: str) -> int:
    value = 0
    for char in column:
        value = value * 26 + ord(char.upper()) - 64
    return value
